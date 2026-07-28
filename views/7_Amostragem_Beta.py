import re

import openpyxl
import streamlit as st
import pandas as pd

from core.amostragem import (
    ORDEM_CRITICAS,
    REGRAS_AMOSTRAGEM,
    _norm,
    carregar_procedimentos_criticos,
    consolidar_por_guia,
    guias_com_proc_critico,
    marcar_amostra,
    renderizar_tabela_guias,
    selecionar_procedimentos_ignorados,
    gerenciar_procedimentos_ignorados,
)
from shared.database import DatabaseManager

st.set_page_config(page_title="Amostragem", page_icon="", layout="wide")

if not st.session_state.get("logado", False):
    st.warning("Você precisa fazer login na página inicial para acessar esta ferramenta.")
    st.stop()

if "db" not in st.session_state:
    st.session_state.db = DatabaseManager()

# Colunas mínimas esperadas na planilha mensal da base IA (mesma que alimenta
# o PowerBI). Nomes normalizados via _norm (maiúsculo, sem acento).
COLUNAS_NECESSARIAS = {
    "NU_ORDEM", "NU_GUIA", "CD_PROCEDIMENTO", "DS_GRUPO", "LIBERACAO",
    "DT_CREATED_AT", "CD_OPERADOR_ATEND",
}

# Operador que indica biometria facial feita (fluxo automático via app);
# qualquer outro operador = análise manual, sem biometria.
OPERADOR_BIOMETRIA = "CONN_APPOD_NEW"

# Colunas mínimas esperadas na planilha mensal de imagem.
COLUNAS_NECESSARIAS_IMAGEM = {"NU_GUIA", "CD_PROCEDIMENTO", "DENTE_INICIAL", "STATUS_PROCED", "NOME_ARQUIVO"}

SEED_PADRAO = 42
_is_admin = st.session_state.get("role_interno") == "Admin"


def _preparar_registros(arquivo) -> tuple[list, str, int]:
    """Lê a planilha mensal e devolve (registros_para_inserir, mes_referencia, total_bruto).

    Lê linha a linha via openpyxl (read_only) em vez de pandas.read_excel —
    a planilha tem ~500 mil linhas e carregar tudo num DataFrame de uma vez
    consome memória demais (gerou MemoryError em máquina com pouca RAM livre).

    Só mantém linhas com LIBERACAO == 'N' (é só isso que a amostragem usa).
    `mes_referencia` é derivado de DT_CREATED_AT (constante por arquivo,
    ex: planilha 'IA 07 2026' tem DT_CREATED_AT = 2026-07-01).
    """
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    aba = "Planilha1" if "Planilha1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]
    linhas = ws.iter_rows(values_only=True)

    header = [_norm(str(c)) for c in next(linhas)]
    idx = {nome: i for i, nome in enumerate(header)}
    faltantes = COLUNAS_NECESSARIAS - set(idx)
    if faltantes:
        raise ValueError(
            "Colunas não encontradas na planilha (aba '" + aba + "'): "
            + ", ".join(sorted(faltantes))
        )

    i_ordem, i_guia = idx["NU_ORDEM"], idx["NU_GUIA"]
    i_cd, i_grupo = idx["CD_PROCEDIMENTO"], idx["DS_GRUPO"]
    i_lib, i_dt = idx["LIBERACAO"], idx["DT_CREATED_AT"]
    i_operador = idx["CD_OPERADOR_ATEND"]

    registros = []
    guias_por_processo = {}  # nu_ordem -> set(nu_guia), inclui S e N — só pra contar o total
    mes_referencia = None
    total_bruto = 0
    for linha in linhas:
        if linha[i_ordem] is None:
            continue
        total_bruto += 1
        if mes_referencia is None and linha[i_dt] is not None:
            dt = linha[i_dt]
            mes_referencia = dt.strftime("%Y-%m") if hasattr(dt, "strftime") else str(dt)[:7]

        nu_ordem = str(int(linha[i_ordem]))
        nu_guia = str(linha[i_guia]).strip()
        guias_por_processo.setdefault(nu_ordem, set()).add(nu_guia)

        liberacao = str(linha[i_lib] or "").strip().upper()
        if liberacao != "N":
            continue

        registros.append({
            "nu_ordem": nu_ordem,
            "nu_guia": nu_guia,
            "cd_procedimento": str(linha[i_cd]).strip(),
            "ds_grupo": str(linha[i_grupo]).strip(),
            "liberacao": "N",
            "mes_referencia": None,
            "total_guias_processo": None,
            "cd_operador_atend": str(linha[i_operador] or "").strip(),
        })

    wb.close()

    if mes_referencia is None:
        raise ValueError("Não foi possível ler DT_CREATED_AT para determinar o mês de referência.")

    for registro in registros:
        registro["mes_referencia"] = mes_referencia
        registro["total_guias_processo"] = len(guias_por_processo[registro["nu_ordem"]])

    return registros, mes_referencia, total_bruto


def _preparar_registros_imagem(arquivo) -> tuple[list, str, int]:
    """Lê a planilha mensal de imagem e devolve (registros, mes_referencia, total_bruto).

    Essa planilha não tem coluna de data — o mês de referência é extraído do
    NOME do arquivo (padrão observado: "MM AAAA <código> - IMAGEM.xlsx", ex:
    "06 2026 4016R - IMAGEM.xlsx" -> mes_referencia = "2026-06").
    """
    match = re.match(r"^\s*(\d{2})\s+(\d{4})", arquivo.name)
    if not match:
        raise ValueError(
            f"Não consegui identificar o mês/ano no nome do arquivo '{arquivo.name}'. "
            "Esperado algo como 'MM AAAA ... .xlsx'."
        )
    mes, ano = match.group(1), match.group(2)
    mes_referencia = f"{ano}-{mes}"

    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    aba = "Planilha1" if "Planilha1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]
    linhas = ws.iter_rows(values_only=True)

    header = [_norm(str(c)) for c in next(linhas)]
    idx = {nome: i for i, nome in enumerate(header)}
    faltantes = COLUNAS_NECESSARIAS_IMAGEM - set(idx)
    if faltantes:
        raise ValueError(
            "Colunas não encontradas na planilha (aba '" + aba + "'): "
            + ", ".join(sorted(faltantes))
        )

    i_guia, i_cd = idx["NU_GUIA"], idx["CD_PROCEDIMENTO"]
    i_dente, i_status, i_arquivo = idx["DENTE_INICIAL"], idx["STATUS_PROCED"], idx["NOME_ARQUIVO"]

    registros = []
    total_bruto = 0
    for linha in linhas:
        if linha[i_guia] is None:
            continue
        total_bruto += 1
        tem_imagem = str(linha[i_arquivo] or "").strip().upper() != "SEM IMAGEM"
        registros.append({
            "nu_guia": str(linha[i_guia]).strip(),
            "cd_procedimento": str(linha[i_cd]).strip(),
            "dente_inicial": str(linha[i_dente]).strip() if linha[i_dente] is not None else None,
            "status_proced": str(linha[i_status]).strip() if linha[i_status] is not None else None,
            "tem_imagem": tem_imagem,
            "mes_referencia": mes_referencia,
        })

    wb.close()
    return registros, mes_referencia, total_bruto


def _guias_para_df(guias: list) -> pd.DataFrame:
    """Converte o retorno do Supabase para o mesmo formato que
    parse_powerbi() produz, pra reaproveitar consolidar_por_guia/marcar_amostra."""
    if not guias:
        return pd.DataFrame()
    return pd.DataFrame({
        "Especialidade": [g["ds_grupo"] for g in guias],
        "CD_PROCEDIMENTO": [g["cd_procedimento"] for g in guias],
        "NU_GUIA": [g["nu_guia"] for g in guias],
        "CD_OPERADOR_ATEND": [g.get("cd_operador_atend", "") for g in guias],
        "Qtde": 1,
    })


# --------------------------------------------------------------------- UI ----

st.title("Amostragem de Guias")
st.caption(
    "Digite o número do processo — as guias com LIBERAÇÃO = N são buscadas "
    "automaticamente na base importada mensalmente. Prestador e percentuais "
    "continuam só no PowerBI."
)

aba_busca, aba_config = st.tabs(["Amostragem", "Configurações"])

with aba_config:
    gerenciar_procedimentos_ignorados(st.session_state.db, key_prefix="amostragem_beta")

    if _is_admin:
        with st.expander("Importar planilha mensal da base IA (Admin)", expanded=False):
            st.caption(
                "Sobe a planilha do mês (mesma que alimenta o PowerBI). Substitui "
                "os dados do mês detectado e mantém só os 2 meses mais recentes na base."
            )
            arquivo = st.file_uploader("Planilha mensal (.xlsx)", type=["xlsx"], key="upload_base_ia")
            if arquivo and st.button("Importar"):
                try:
                    with st.spinner("Lendo e importando (pode levar alguns minutos)..."):
                        registros, mes_referencia, total_bruto = _preparar_registros(arquivo)
                        if not registros:
                            st.warning("Nenhuma linha com LIBERAÇÃO = N encontrada nesta planilha.")
                        else:
                            total_inserido = st.session_state.db.importar_base_ia(registros, mes_referencia)
                            st.success(
                                f"Mês {mes_referencia}: {total_inserido} de {total_bruto} linha(s) "
                                f"(LIBERAÇÃO = N) importadas com sucesso."
                            )
                except Exception as erro:
                    st.error(f"Falha na importação: {erro}")

        with st.expander("Importar planilha mensal de imagem (Admin)", expanded=False):
            st.caption(
                "Sobe a planilha de imagem do mês (colunas NU_GUIA, CD_PROCEDIMENTO, "
                "DENTE_INICIAL, STATUS_PROCED, NOME_ARQUIVO). O mês de referência é lido "
                "do nome do arquivo (ex: '06 2026 ... .xlsx' → junho/2026)."
            )
            arquivo_imagem = st.file_uploader("Planilha de imagem (.xlsx)", type=["xlsx"], key="upload_base_imagem")
            if arquivo_imagem and st.button("Importar imagem", key="btn_importar_imagem"):
                try:
                    with st.spinner("Lendo e importando (pode levar alguns minutos)..."):
                        registros_img, mes_referencia_img, total_bruto_img = _preparar_registros_imagem(arquivo_imagem)
                        if not registros_img:
                            st.warning("Nenhuma linha encontrada nesta planilha.")
                        else:
                            total_inserido_img = st.session_state.db.importar_base_imagem(registros_img, mes_referencia_img)
                            st.success(
                                f"Mês {mes_referencia_img}: {total_inserido_img} de {total_bruto_img} "
                                f"linha(s) importadas com sucesso."
                            )
                except Exception as erro:
                    st.error(f"Falha na importação: {erro}")

with aba_busca:
    processo_digitado = st.text_input("Número do processo", placeholder="Ex: 8202650447")
    buscar = st.button("Buscar guias")

    if buscar:
        st.session_state["_amostragem_beta_processo"] = processo_digitado.strip()

    processo_ativo = st.session_state.get("_amostragem_beta_processo", "")

    if not processo_ativo:
        st.info("Digite o número do processo e clique em Buscar guias.")
        st.stop()

    guias = st.session_state.db.buscar_guias_ia_por_processo(processo_ativo)
    df = _guias_para_df(guias)

    if df.empty:
        st.warning(
            f"Nenhuma guia com LIBERAÇÃO = N encontrada para o processo "
            f"'{processo_ativo}' na base importada. Confira o número ou se o "
            f"mês do processo ainda está entre os 2 meses mantidos na base."
        )
        st.stop()

    total_guias_processo = guias[0].get("total_guias_processo") if guias else None
    texto_total_guias = f" — {total_guias_processo} guia(s) no total do processo" if total_guias_processo else ""
    st.success(f"Processo {processo_ativo}: {len(df)} item(ns) sem liberação pela IA{texto_total_guias}.")

    # Biometria por guia: computado do df ANTES do filtro de procedimentos
    # ignorados (é atributo de quem atendeu, não depende de quais
    # procedimentos entram ou não na análise). Guarda (qtd_com_biometria,
    # qtd_total_itens, qtd_com_operador_gravado) — o check só aparece quando
    # os dois primeiros batem (100%); se nenhum item tiver operador gravado
    # (guias importadas antes desta coluna existir), a célula fica em branco
    # em vez de mostrar "0/N" (que passaria a entender errada de "sem
    # biometria" quando na real é "dado ainda não disponível").
    def _biometria_guia(serie):
        valores = [str(v).strip() for v in serie]
        n_total = len(valores)
        n_bio = sum(v == OPERADOR_BIOMETRIA for v in valores)
        n_com_dado = sum(1 for v in valores if v)
        return (n_bio, n_total, n_com_dado)

    biometria_por_guia = (
        df.groupby("NU_GUIA")["CD_OPERADOR_ATEND"]
        .apply(_biometria_guia)
        .to_dict()
    )

    # Imagem por guia: vem de uma base separada (planilha própria, cadência
    # própria), cruzada por NU_GUIA. Guarda (qtd_com_imagem, qtd_total) —
    # aqui não existe "dado legado nulo" como na biometria (tem_imagem é
    # sempre um booleano real desde a importação), então guia ausente do
    # dict = sem nenhum registro de imagem ainda (célula em branco).
    imagem_registros = st.session_state.db.buscar_imagem_por_guias(df["NU_GUIA"].unique().tolist())
    imagem_por_guia = {}
    for reg in imagem_registros:
        n_ok, n_total = imagem_por_guia.get(reg["nu_guia"], (0, 0))
        n_total += 1
        if reg.get("tem_imagem"):
            n_ok += 1
        imagem_por_guia[reg["nu_guia"]] = (n_ok, n_total)

    # --- Filtros: procedimentos ignorados + Biometria/Imagem, agrupados num
    # único painel recolhido por padrão (não competem com o resultado abaixo).
    with st.expander("Filtros", expanded=False):
        # Procedimentos que não precisam ser analisados nesta guia (ex.:
        # coroas provisórias). O procedimento some da contagem e do sorteio;
        # a guia continua listada mesmo que fique sem nenhum procedimento
        # restante. A seleção pode ser salva como padrão (por especialidade),
        # aplicado automaticamente nas próximas análises.
        codigos_excluidos = selecionar_procedimentos_ignorados(
            df, st.session_state.db, key_prefix=f"amostragem_beta_{processo_ativo}"
        )

        st.divider()
        col_filtro_bio, col_filtro_img = st.columns(2)
        with col_filtro_bio:
            st.caption("Biometria")
            filtro_biometria = st.segmented_control(
                "Biometria", ["Todos", "Com", "Sem"],
                default="Todos", key=f"filtro_biometria_{processo_ativo}",
                label_visibility="collapsed",
            )
        with col_filtro_img:
            st.caption("Imagem")
            filtro_imagem = st.segmented_control(
                "Imagem", ["Todos", "Com", "Sem"],
                default="Todos", key=f"filtro_imagem_{processo_ativo}",
                label_visibility="collapsed",
            )

    todas_guias = df[["Especialidade", "NU_GUIA"]].drop_duplicates()
    df = df[~df["CD_PROCEDIMENTO"].isin(codigos_excluidos)] if codigos_excluidos else df

    df_guias = consolidar_por_guia(df)
    if codigos_excluidos:
        df_guias = todas_guias.merge(df_guias, on=["Especialidade", "NU_GUIA"], how="left")
        df_guias["Procedimentos"] = df_guias["Procedimentos"].fillna("")
        df_guias["Qtde_procs"] = df_guias["Qtde_procs"].fillna(0).astype(int)
        # Prótese: guia some da lista inteira se sobrou sem nenhum
        # procedimento (só tinha o(s) ignorado(s)) — nas demais
        # especialidades a guia continua aparecendo mesmo vazia.
        vazia_protese = (df_guias["Especialidade"].apply(_norm) == "PROTESE") & (df_guias["Qtde_procs"] == 0)
        df_guias = df_guias[~vazia_protese]
        df_guias = df_guias.sort_values(["Especialidade", "Procedimentos", "NU_GUIA"]).reset_index(drop=True)

    # --- Filtros de Biometria e Imagem ---
    # "Sem dado" (guia ainda não aparece na base de biometria/imagem) conta
    # como "não 100%" no filtro "Sem" — é o lado mais seguro pra auditoria
    # (não confirmado = trata como pendência).
    def _guia_100pct(info):
        if not info:
            return None
        n_ok, n_total = info[0], info[1]
        n_com_dado = info[2] if len(info) > 2 else n_total
        if n_total == 0 or n_com_dado == 0:
            return None
        return n_ok == n_total

    def _aplicar_filtro_guia(df_in, mapa, filtro):
        if not filtro or filtro == "Todos":
            return df_in
        quer_com = filtro == "Com"

        def _passa(guia):
            resultado = _guia_100pct(mapa.get(str(guia)))
            if resultado is None:
                return not quer_com
            return resultado == quer_com

        return df_in[df_in["NU_GUIA"].apply(_passa)]

    df_guias = _aplicar_filtro_guia(df_guias, biometria_por_guia, filtro_biometria)
    df_guias = _aplicar_filtro_guia(df_guias, imagem_por_guia, filtro_imagem)

    if df_guias.empty:
        st.info("Nenhuma guia bate com os filtros selecionados.")
        st.stop()

    # Mantém df (nível de item) em sincronia com as guias que sobraram após
    # os filtros — senão os totais de procedimentos no Resumo contariam
    # guias que os filtros já tiraram da lista.
    df = df[df["NU_GUIA"].isin(df_guias["NU_GUIA"])]

    guias_vistas = st.session_state.db.buscar_guias_vistas(df_guias["NU_GUIA"].unique().tolist())

    # Procedimentos cadastrados como críticos (tabela_procedimentos.critico) —
    # só importam pras especialidades fora de REGRAS_AMOSTRAGEM (Periodontia,
    # Odontopediatria, Radiologia Especial etc.): se uma guia dessas
    # especialidades tiver um desses procedimentos, ela sobe na lista e entra
    # garantida na "Sugestão de amostra", já que são casos raros que o
    # auditor corre risco de não perceber se ficarem escondidos lá embaixo.
    procedimentos_criticos = carregar_procedimentos_criticos()

    especialidades = df_guias["Especialidade"].unique().tolist()

    # Pra cada especialidade fora de REGRAS_AMOSTRAGEM, marca se tem pelo
    # menos uma guia com procedimento crítico nesta análise — decide tanto a
    # posição na lista quanto o conteúdo da "Sugestão de amostra" abaixo.
    especialidade_tem_critico = {}
    for esp in especialidades:
        if _norm(esp) in REGRAS_AMOSTRAGEM:
            continue
        df_esp_guias_check = df_guias[df_guias["Especialidade"] == esp]
        especialidade_tem_critico[esp] = not guias_com_proc_critico(df_esp_guias_check, procedimentos_criticos).empty

    def _peso_ordenacao(e):
        norm = _norm(e)
        # Procedimento crítico presente = topo absoluto, acima até das
        # especialidades já críticas — é justamente o caso raro que corre
        # risco de passar despercebido, então precisa ser o mais visível.
        if especialidade_tem_critico.get(e):
            return (0, norm)
        if norm in ORDEM_CRITICAS:
            return (1, ORDEM_CRITICAS.index(norm))
        return (2, norm)

    especialidades.sort(key=_peso_ordenacao)

    # --- Resumo ---
    resumo = []
    for esp in especialidades:
        df_esp_total = df[df["Especialidade"] == esp]
        df_esp_guias = df_guias[df_guias["Especialidade"] == esp].reset_index(drop=True)
        total_procs = int(df_esp_total["Qtde"].sum())
        total_guias = len(df_esp_guias)

        if _norm(esp) in REGRAS_AMOSTRAGEM:
            n_sugerido = len(marcar_amostra(df_esp_guias, esp, df_esp_total, seed=SEED_PADRAO))
        else:
            n_sugerido = len(guias_com_proc_critico(df_esp_guias, procedimentos_criticos))

        resumo.append({
            "Especialidade": esp,
            "Guias únicas": total_guias,
            "Total de procs": total_procs,
            "Amostra sugerida": n_sugerido,
        })

    st.markdown("### Resumo")
    st.dataframe(pd.DataFrame(resumo), use_container_width=True, hide_index=True)

    # --- Detalhamento ---
    st.markdown("### Detalhamento por especialidade")
    st.caption("Clique no número da guia para copiar.")

    for esp in especialidades:
        df_esp_total = df[df["Especialidade"] == esp]
        df_esp_guias = df_guias[df_guias["Especialidade"] == esp].reset_index(drop=True)
        total_procs = int(df_esp_total["Qtde"].sum())
        total_guias = len(df_esp_guias)

        def _tabela_completa():
            renderizar_tabela_guias(
                df_esp_guias, esp, objetivo=total_guias,
                guias_vistas=guias_vistas, biometria_por_guia=biometria_por_guia,
                imagem_por_guia=imagem_por_guia,
            )

        df_amostra_especial = None
        titulo_amostra = None

        if _norm(esp) in REGRAS_AMOSTRAGEM:
            df_amostra = marcar_amostra(df_esp_guias, esp, df_esp_total, seed=SEED_PADRAO)
            # Prótese sempre audita 100% das guias (regra "todas") — a
            # "Sugestão de amostra" ficaria idêntica à "Tabela completa",
            # então não ganha aba própria pra essa especialidade (as demais
            # com regra "todas", como Implante, continuam mostrando normalmente).
            if _norm(esp) != "PROTESE":
                df_amostra_especial = df_amostra.drop(columns=["Motivo"], errors="ignore")
                titulo_amostra = f"Sugestão de amostra ({len(df_amostra_especial)})"
        elif especialidade_tem_critico.get(esp):
            # Especialidade fora das regras de amostragem, mas com
            # procedimento crítico presente: a "Sugestão de amostra" mostra
            # só as guias com esse procedimento, não as 100% da especialidade.
            df_amostra_especial = guias_com_proc_critico(df_esp_guias, procedimentos_criticos)
            titulo_amostra = f"Sugestão de amostra ({len(df_amostra_especial)} — proc. crítico)"
        # Sem regra de amostragem e sem procedimento crítico presente: sem
        # aba de "Sugestão de amostra" (hoje mostraria 100% das guias, igual
        # à Tabela completa, sem utilidade nenhuma).

        with st.expander(f"{esp} — {total_guias} guia(s), {total_procs} proc(s)", expanded=False):
            if df_amostra_especial is not None:
                tab_completa, tab_amostra = st.tabs([f"Tabela completa ({total_guias})", titulo_amostra])
                with tab_completa:
                    _tabela_completa()
                with tab_amostra:
                    renderizar_tabela_guias(
                        df_amostra_especial, esp, objetivo=len(df_amostra_especial),
                        guias_vistas=guias_vistas, biometria_por_guia=biometria_por_guia,
                        imagem_por_guia=imagem_por_guia,
                    )
            else:
                _tabela_completa()
