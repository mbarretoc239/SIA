import html
import json
import re
import unicodedata

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from core.requisitos_procedimentos import REQUISITOS_POR_PROCEDIMENTO

# Tradução das siglas curtas de core/requisitos_procedimentos.py -- usado só
# no tooltip (hover) da coluna REQUISITOS da Amostragem, ver
# renderizar_tabela_guias._decodificar_sigla_requisito.
_DECODIFICA_SIGLA_REQUISITO = {
    "P": "Película",
    "RX": "Radiografia",
    "RXF": "Radiografia Final",
    "RXI": "Radiografia Inicial",
    "RXIF": "Radiografia Inicial e Final",
    "TC": "Termo de Consentimento e Ciência",
    "DOC": "Documentação",
    "E": "Encaminhamento do CD assistente",
    "FF": "Fotografia Final",
    "FI": "Fotografia Inicial",
    "FIF": "Fotografia Inicial e Final",
    "F": "Fotografia",
    "L": "Laudo Técnico",
    "-": "Nenhum requisito específico",
}

# Quantidade de imagens/documentos esperada por sigla de requisito -- usado
# pra corrigir o denominador do badge IMAGEM na Amostragem, que antes vinha
# só de quantas linhas a planilha 4016R trouxe pra guia (podendo faltar
# linha pro que o requisito realmente pede). *IF (inicial+final) conta 2;
# DOC fica de fora (quantidade variável, às vezes mais de 2 documentos,
# não dá pra fixar um número) -- ver conversa 2026-08-20.
_QTD_IMAGEM_POR_SIGLA = {
    "-": 0, "DOC": 0,
    "P": 1, "RX": 1, "RXF": 1, "RXI": 1, "TC": 1, "E": 1, "L": 1,
    "FF": 1, "FI": 1, "F": 1,
    "RXIF": 2, "FIF": 2,
}


def _qtd_imagens_esperada(requisito: str) -> int:
    """Quantidade de imagens/documentos esperada pra um requisito (pode ter
    '+' = precisa dos dois, soma; '/' = precisa de só um dos dois, usa o
    menor valor entre as alternativas, que é o mínimo que já satisfaz)."""
    grupos_ou = requisito.split("/")
    valores = []
    for grupo in grupos_ou:
        siglas = grupo.split("+")
        valores.append(sum(_QTD_IMAGEM_POR_SIGLA.get(s.strip(), 0) for s in siglas))
    return min(valores) if valores else 0


def calcular_imagens_esperadas_guia(procedimentos_str) -> int:
    """Soma a quantidade de imagens/documentos esperada pra guia, por
    procedimento (cada procedimento pede as suas, não compartilha com os
    outros da mesma guia). Procedimento sem requisito mapeado não soma
    nada -- não inventa exigência pro que não temos dado."""
    codigos = [c.strip() for c in str(procedimentos_str).split(",") if c.strip()]
    return sum(
        _qtd_imagens_esperada(REQUISITOS_POR_PROCEDIMENTO[cod])
        for cod in codigos if cod in REQUISITOS_POR_PROCEDIMENTO
    )


# Colunas mínimas esperadas na planilha mensal da base IA (mesma que alimenta
# o PowerBI). Nomes normalizados via _norm (maiúsculo, sem acento).
COLUNAS_NECESSARIAS_BASE_IA = {
    "NU_ORDEM", "NU_GUIA", "CD_PROCEDIMENTO", "DS_GRUPO", "LIBERACAO",
    "DT_PRODUCAO", "CD_OPERADOR_ATEND",
}

# Colunas mínimas esperadas na planilha mensal de imagem. CD_PROCEDIMENTO e
# STATUS_PROCED não são exigidas: o layout mudou (planilha atual não traz
# nem uma nem outra) e nenhuma das duas é lida hoje na tela — só tem_imagem
# (via NOME_ARQUIVO) e nu_guia são realmente usados no cruzamento.
COLUNAS_NECESSARIAS_IMAGEM = {"NU_GUIA", "DENTE_INICIAL", "NOME_ARQUIVO"}


# Especialidades tratadas como críticas (maior valor/risco, normalmente com
# exigência de imagem). Periodontia NÃO está mais aqui — passou a ser tratada
# como as demais especialidades "comuns", sujeita só à regra de procedimento
# crítico abaixo (ver PROCEDIMENTOS_CRITICOS_ESPECIALIDADE_COMUM).
#
# Fallback usado só se a tabela amostragem_regras_amostra (Supabase) estiver
# vazia ou inacessível -- a fonte de verdade normal é o banco, editável por
# Admin/Gestor em Configurações > Regras de Amostragem (ver
# carregar_regras_amostragem_cache abaixo e DatabaseManager.carregar_regras_amostragem).
REGRAS_AMOSTRAGEM_PADRAO = {
    "IMPLANTE": {"tipo": "todas"},
    "PROTESE": {"tipo": "todas"},
    "PROTESE ESPECIAL": {"tipo": "todas"},
    "CIRURGIA": {"tipo": "percentual", "pct": 0.30, "minimo_procs": 10, "minimo_amostra": 5},
    "ENDODONTIA": {"tipo": "percentual", "pct": 0.50, "minimo_procs": 10},
}

ORDEM_CRITICAS_PADRAO = [
    "IMPLANTE",
    "PROTESE",
    "PROTESE ESPECIAL",
    "CIRURGIA",
    "ENDODONTIA",
]


@st.cache_data(ttl=60)
def carregar_regras_amostragem_cache() -> tuple[dict, list]:
    """(REGRAS_AMOSTRAGEM, ORDEM_CRITICAS) carregados de
    amostragem_regras_amostra (Supabase), na mesma forma que os dicts/listas
    hardcoded de antes -- só ignora linhas com ativo=false e ordena por
    'ordem'. Cai pro fallback _PADRAO acima se a tabela vier vazia (nunca
    configurada ainda) ou se a consulta falhar, pra nunca deixar a
    Amostragem sem nenhuma regra por causa de um problema de rede."""
    from shared.database import DatabaseManager
    linhas = DatabaseManager().carregar_regras_amostragem()
    ativas = [l for l in linhas if l.get("ativo", True)]
    if not ativas:
        return dict(REGRAS_AMOSTRAGEM_PADRAO), list(ORDEM_CRITICAS_PADRAO)

    regras = {}
    for l in sorted(ativas, key=lambda l: l.get("ordem", 100)):
        esp = _norm(l["especialidade"])
        if l["tipo"] == "todas":
            regras[esp] = {"tipo": "todas"}
        else:
            regra = {"tipo": "percentual", "pct": float(l["pct"]), "minimo_procs": int(l["minimo_procs"] or 0)}
            if l.get("minimo_amostra"):
                regra["minimo_amostra"] = int(l["minimo_amostra"])
            regras[esp] = regra
    ordem_criticas = list(regras.keys())
    return regras, ordem_criticas

# Procedimentos de baixa prioridade pra amostra, por especialidade — os mais
# recorrentes/baratos, que passam pelo sorteio normal. Guias que contêm
# QUALQUER procedimento fora desta lista são de alta prioridade e entram na
# amostra automaticamente (sem depender do sorteio).
#
# Motivação: certos procedimentos são raros e/ou custosos e não podem cair
# fora da amostra por acaso do sorteio (ex.: exodontia de incluso na
# CIRURGIA). Deixá-los "sempre auditar" garante cobertura.
#
# Não confundir com "especialidade crítica" (ORDEM_CRITICAS/REGRAS_AMOSTRAGEM
# acima) nem com "procedimento crítico" (PROCEDIMENTOS_CRITICOS_ESPECIALIDADE_
# COMUM abaixo) — são três conceitos independentes: este aqui é só sobre
# prioridade de procedimento dentro da amostra, DENTRO de uma especialidade
# já crítica.
PROCS_PRIORIDADE_NORMAL = {
    "CIRURGIA": {"5010", "5030", "5031"},
    "ENDODONTIA": {"2015", "2025", "2035"},
}

# Especialidades que NÃO estão em REGRAS_AMOSTRAGEM (ex.: Periodontia,
# Odontopediatria, Radiologia Especial...) normalmente não têm seção de
# "Sugestão de amostra" própria — hoje mostrariam 100% das guias, igual à
# Tabela completa, o que não ajuda em nada.
#
# Exceção: se alguma guia dessa especialidade tiver um procedimento marcado
# como crítico em `tabela_procedimentos.critico` (cadastro oficial, carregado
# por carregar_procedimentos_criticos), a especialidade sobe pro topo da
# lista de "Detalhamento por especialidade" e a "Sugestão de amostra" passa a
# mostrar só as guias com esse procedimento — são casos raros que o auditor
# corre risco de não perceber se ficarem escondidos lá embaixo.
@st.cache_data(ttl=300)
def carregar_procedimentos_criticos() -> set:
    from shared.database import DatabaseManager
    db = DatabaseManager()
    url = f"{db.supabase_url}/rest/v1/tabela_procedimentos?select=codigo_tuss&critico=eq.true"
    import requests
    r = requests.get(url, headers=db.headers)
    if not r.ok:
        return set()
    return {str(row["codigo_tuss"]).strip() for row in r.json()}


def _norm(texto: str) -> str:
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    return sem_acento.strip().upper()


def preparar_registros_base_ia(arquivo) -> tuple[list, str, int]:
    """Lê a planilha mensal da base IA e devolve (registros_para_inserir, mes_referencia, total_bruto).

    Lê linha a linha via openpyxl (read_only) em vez de pandas.read_excel —
    a planilha tem centenas de milhares de linhas e carregar tudo num
    DataFrame de uma vez consome memória demais (gerou MemoryError em
    máquina com pouca RAM livre).

    Traz TODAS as linhas (liberadas e não liberadas) -- antes só trazia
    LIBERACAO == 'N', mas o "% Biometria" da lista de processos (Amostragem)
    precisa da base inteira do processo, não só das guias pendentes de
    revisão. As telas que só devem enxergar o pendente (tabela de guias da
    Amostragem, contagem de especialidades/críticas) filtram
    `liberacao = 'N'` explicitamente do lado do SQL agora -- não dependem
    mais implicitamente de "só existe N na tabela".
    `mes_referencia` é derivado de DT_PRODUCAO (constante por arquivo,
    ex: planilha 'IA 07 2026' tem DT_PRODUCAO = 2026-07-01).
    """
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    aba = "Planilha1" if "Planilha1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]
    linhas = ws.iter_rows(values_only=True)

    header = [_norm(str(c)) for c in next(linhas)]
    idx = {nome: i for i, nome in enumerate(header)}
    faltantes = COLUNAS_NECESSARIAS_BASE_IA - set(idx)
    if faltantes:
        raise ValueError(
            "Colunas não encontradas na planilha (aba '" + aba + "'): "
            + ", ".join(sorted(faltantes))
        )

    i_ordem, i_guia = idx["NU_ORDEM"], idx["NU_GUIA"]
    i_cd, i_grupo = idx["CD_PROCEDIMENTO"], idx["DS_GRUPO"]
    i_lib, i_dt = idx["LIBERACAO"], idx["DT_PRODUCAO"]
    i_operador = idx["CD_OPERADOR_ATEND"]

    registros = []
    guias_por_processo = {}  # nu_ordem -> set(nu_guia), inclui S e N — só pra contar o total
    mes_referencia = None
    total_bruto = 0
    for linha in linhas:
        if linha[i_ordem] is None:
            continue
        total_bruto += 1
        if mes_referencia is None and linha[i_dt] is not None:
            dt = linha[i_dt]
            mes_referencia = dt.strftime("%Y-%m") if hasattr(dt, "strftime") else str(dt)[:7]

        nu_ordem = str(int(linha[i_ordem]))
        nu_guia = str(linha[i_guia]).strip()
        guias_por_processo.setdefault(nu_ordem, set()).add(nu_guia)

        liberacao = str(linha[i_lib] or "").strip().upper()

        registros.append({
            "nu_ordem": nu_ordem,
            "nu_guia": nu_guia,
            "cd_procedimento": str(linha[i_cd]).strip(),
            "ds_grupo": str(linha[i_grupo]).strip(),
            "liberacao": liberacao,
            "mes_referencia": None,
            "total_guias_processo": None,
            "cd_operador_atend": str(linha[i_operador] or "").strip(),
        })

    wb.close()

    if mes_referencia is None:
        raise ValueError("Não foi possível ler DT_PRODUCAO para determinar o mês de referência.")

    for registro in registros:
        registro["mes_referencia"] = mes_referencia
        registro["total_guias_processo"] = len(guias_por_processo[registro["nu_ordem"]])

    return registros, mes_referencia, total_bruto


def preparar_registros_imagem(arquivo) -> tuple[list, str, int]:
    """Lê a planilha mensal de imagem e devolve (registros, mes_referencia, total_bruto).

    Essa planilha não tem coluna de data — o mês de referência é extraído do
    NOME do arquivo (padrão observado: "MM AAAA <código> - IMAGEM.xlsx", ex:
    "06 2026 4016R - IMAGEM.xlsx" -> mes_referencia = "2026-06").
    """
    match = re.match(r"^\s*(\d{2})\s+(\d{4})", arquivo.name)
    if not match:
        raise ValueError(
            f"Não consegui identificar o mês/ano no nome do arquivo '{arquivo.name}'. "
            "Esperado algo como 'MM AAAA ... .xlsx'."
        )
    mes, ano = match.group(1), match.group(2)
    mes_referencia = f"{ano}-{mes}"

    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    aba = "Planilha1" if "Planilha1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]
    linhas = ws.iter_rows(values_only=True)

    header = [_norm(str(c)) for c in next(linhas)]
    idx = {nome: i for i, nome in enumerate(header)}
    faltantes = COLUNAS_NECESSARIAS_IMAGEM - set(idx)
    if faltantes:
        raise ValueError(
            "Colunas não encontradas na planilha (aba '" + aba + "'): "
            + ", ".join(sorted(faltantes))
        )

    i_guia, i_dente, i_arquivo = idx["NU_GUIA"], idx["DENTE_INICIAL"], idx["NOME_ARQUIVO"]
    # Opcionais: presentes no layout antigo, ausentes no atual (que trouxe
    # NOME_PROCEDIMENTO/CODIGO no lugar, sem serem substitutos diretos).
    i_cd = idx.get("CD_PROCEDIMENTO")
    i_status = idx.get("STATUS_PROCED")

    registros = []
    total_bruto = 0
    for linha in linhas:
        if linha[i_guia] is None:
            continue
        total_bruto += 1
        tem_imagem = str(linha[i_arquivo] or "").strip().upper() != "SEM IMAGEM"
        registros.append({
            "nu_guia": str(linha[i_guia]).strip(),
            # cd_procedimento é NOT NULL na tabela; "" quando a coluna não existe na planilha.
            "cd_procedimento": str(linha[i_cd]).strip() if i_cd is not None and linha[i_cd] is not None else "",
            "dente_inicial": str(linha[i_dente]).strip() if linha[i_dente] is not None else None,
            "status_proced": str(linha[i_status]).strip() if i_status is not None and linha[i_status] is not None else None,
            "tem_imagem": tem_imagem,
            "mes_referencia": mes_referencia,
        })

    wb.close()
    return registros, mes_referencia, total_bruto


# Especialidades (DS_GRUPO) conhecidas na base — usado só para popular o
# seletor do painel de gerenciamento; não limita o que pode ser sorteado.
ESPECIALIDADES_CONHECIDAS = [
    "CIRURGIA",
    "CONSULTA INICIAL",
    "ENDODONTIA",
    "ESTETICA",
    "EXAMES COMPLEMENTARES",
    "IMPLANTE",
    "ODONTOLOGIA CLINICA",
    "ODONTOPEDIATRIA",
    "ORTODONTIA",
    "PERIODONTIA",
    "PREVENCAO",
    "PROTESE",
    "PROTESE ESPECIAL",
    "RADIOLOGIA",
    "RADIOLOGIA ESPECIAL",
    "URGENCIA/EMERGENCIA",
]


def gerenciar_procedimentos_ignorados(db, key_prefix: str):
    """Painel para adicionar/remover procedimentos ignorados buscando em todo
    o catálogo (services.relatorio_5302.glosa_matcher.carregar_mapa_procedimentos),
    não só os presentes numa análise já carregada. Complementa
    selecionar_procedimentos_ignorados, que só oferece o que já está no
    dataset da análise atual."""
    from services.relatorio_5302.glosa_matcher import carregar_mapa_procedimentos

    with st.expander("Gerenciar procedimentos ignorados (todas as especialidades)"):
        salvos = db.carregar_procs_ignorados()

        if any(salvos.values()):
            st.markdown("**Já salvos**")
            for esp in sorted(k for k, v in salvos.items() if v):
                codigos = sorted(salvos[esp])
                col_esp, col_codigos, col_rem = st.columns([2, 4, 3])
                with col_esp:
                    st.markdown(f"**{esp}**")
                with col_codigos:
                    st.caption(", ".join(codigos))
                with col_rem:
                    remover_selecionados = st.multiselect(
                        "Remover",
                        options=codigos,
                        key=f"{key_prefix}_gerenciar_remover_{esp}",
                        label_visibility="collapsed",
                        placeholder="Remover código...",
                    )
                    if remover_selecionados and st.button(
                        "Remover", key=f"{key_prefix}_gerenciar_btn_remover_{esp}", use_container_width=True
                    ):
                        pares = [(esp, cod) for cod in remover_selecionados]
                        if db.remover_procs_ignorados(pares):
                            st.toast(f"Removido(s) de {esp}.")
                            st.rerun()
                        else:
                            st.error("Erro ao remover.")
            st.divider()

        st.markdown("**Adicionar novo**")
        mapa_procedimentos = carregar_mapa_procedimentos()
        opcoes_todas = {
            f"{cod} - {desc}": cod
            for cod, desc in sorted(mapa_procedimentos.items(), key=lambda x: x[1])
        }
        col_esp_novo, col_proc_novo = st.columns([1, 3])
        with col_esp_novo:
            especialidade_nova = st.selectbox(
                "Especialidade",
                ESPECIALIDADES_CONHECIDAS,
                key=f"{key_prefix}_gerenciar_esp_nova",
                width=300,
            )
        with col_proc_novo:
            labels_novos = st.multiselect(
                "Procedimentos a ignorar",
                options=sorted(opcoes_todas.keys()),
                key=f"{key_prefix}_gerenciar_proc_novo",
                placeholder="Busque por código ou descrição...",
            )
        if st.button("Adicionar à lista", key=f"{key_prefix}_gerenciar_btn_add"):
            if not labels_novos:
                st.warning("Selecione ao menos um procedimento.")
            else:
                pares = [(especialidade_nova, opcoes_todas[lbl]) for lbl in labels_novos]
                with st.spinner("Salvando..."):
                    sucesso = db.salvar_procs_ignorados(pares)
                if sucesso:
                    st.toast(f"Adicionado(s) a {especialidade_nova}.")
                    st.rerun()
                else:
                    st.error("Erro ao adicionar.")


def selecionar_procedimentos_ignorados(df: pd.DataFrame, db, key_prefix: str) -> set:
    """Renderiza o multiselect "Ignorar procedimentos nesta análise" + botão
    para salvar a seleção como padrão (por especialidade, persistido em
    amostragem_procs_ignorados). A seleção salva vira default automático nas
    próximas análises, mas pode ser ajustada só-nesta-sessão sem salvar.

    `key_prefix` deve variar por dataset (ex.: versão do texto colado, ou o
    processo buscado) para o multiselect resetar corretamente ao trocar de
    entrada — mesmo padrão já usado no key do texto/processo.

    Retorna o set de códigos selecionados AGORA (salvos ou não), pronto para
    filtrar o `df`.
    """
    from services.relatorio_5302.glosa_matcher import carregar_mapa_procedimentos

    mapa_procedimentos = carregar_mapa_procedimentos()

    # Primeira especialidade em que cada código aparece neste dataset —
    # usado pra saber em qual especialidade salvar/remover o código.
    cod_para_especialidade = {}
    for _, row in df[["CD_PROCEDIMENTO", "Especialidade"]].drop_duplicates().iterrows():
        cod_para_especialidade.setdefault(row["CD_PROCEDIMENTO"], row["Especialidade"])

    codigos_presentes = sorted(cod_para_especialidade.keys())
    opcoes = {
        f"{cod} - {mapa_procedimentos.get(cod, 'descrição não encontrada')}": cod
        for cod in codigos_presentes
    }

    salvos = db.carregar_procs_ignorados()  # {especialidade: set(codigos)}
    default_labels = [
        lbl for lbl, cod in opcoes.items()
        if cod in salvos.get(cod_para_especialidade[cod], set())
    ]

    col_multi, col_salvar = st.columns([5, 1.2])
    with col_multi:
        labels_selecionados = st.multiselect(
            "Ignorar procedimentos nesta análise",
            options=sorted(opcoes.keys()),
            default=default_labels,
            key=f"{key_prefix}_procs_ignorados",
            help=(
                "Selecionados aqui não entram na contagem nem no sorteio. "
                "Clique em \"Salvar como padrão\" para aplicar automaticamente "
                "nas próximas análises, sem precisar marcar de novo."
            ),
        )
    codigos_selecionados = {opcoes[lbl] for lbl in labels_selecionados}

    with col_salvar:
        st.write("")
        if st.button("Salvar como padrão", key=f"{key_prefix}_salvar_procs", use_container_width=True):
            pares_para_salvar = [
                (cod_para_especialidade[cod], cod)
                for cod in codigos_selecionados
                if cod not in salvos.get(cod_para_especialidade[cod], set())
            ]
            pares_para_remover = [
                (esp, cod)
                for cod, esp in cod_para_especialidade.items()
                if cod in salvos.get(esp, set()) and cod not in codigos_selecionados
            ]
            ok = True
            if pares_para_salvar:
                ok = ok and db.salvar_procs_ignorados(pares_para_salvar)
            if pares_para_remover:
                ok = ok and db.remover_procs_ignorados(pares_para_remover)
            if ok:
                st.toast("Padrão salvo — aplicado automaticamente nas próximas análises.")
            else:
                st.error("Erro ao salvar o padrão.")

    return codigos_selecionados


def consolidar_por_guia(df: pd.DataFrame) -> pd.DataFrame:
    """Agrupa por (Especialidade, NU_GUIA), juntando procedimentos.

    Ordenado por Procedimentos (código) pra facilitar o escaneio da
    tabela — sem isso as guias apareciam em qualquer ordem, misturando
    códigos diferentes ao longo da lista.

    'Procedimentos' guarda só os códigos puros (usado pela lógica de
    prioridade/crítico/requisitos, que não deve lidar com sufixo de
    quantidade). 'Procedimentos_qtd' guarda a lista [(código, quantidade)]
    na mesma ordem -- usada só na hora de renderizar a tabela, pra marcar
    com "×N" o código que se repete na guia (ex.: 2x do mesmo procedimento
    somava escondido no total da especialidade sem aparecer em lugar
    nenhum)."""
    if df.empty:
        return df
    base = (
        df.groupby(["Especialidade", "NU_GUIA"], sort=False)
        .agg(
            Procedimentos=("CD_PROCEDIMENTO", lambda s: ", ".join(sorted(set(s)))),
            Qtde_procs=("Qtde", "sum"),
        )
        .reset_index()
    )

    qtd_por_codigo = (
        df.groupby(["Especialidade", "NU_GUIA", "CD_PROCEDIMENTO"])["Qtde"]
        .sum()
        .reset_index()
        .sort_values("CD_PROCEDIMENTO")
    )
    procs_qtd = (
        qtd_por_codigo.groupby(["Especialidade", "NU_GUIA"])
        .apply(lambda g: list(zip(g["CD_PROCEDIMENTO"], g["Qtde"].astype(int))), include_groups=False)
        .rename("Procedimentos_qtd")
    )
    base = base.merge(procs_qtd, on=["Especialidade", "NU_GUIA"], how="left")

    return (
        base.sort_values(["Especialidade", "Procedimentos", "NU_GUIA"])
        .reset_index(drop=True)
    )


def calcular_amostra(especialidade: str, total_procs: int, total_guias: int, regras: dict = None):
    """Retorna (n_amostra, descricao_da_regra)."""
    if regras is None:
        regras, _ = carregar_regras_amostragem_cache()
    regra = regras.get(_norm(especialidade))
    if not regra:
        return total_guias, "Não-crítica — sem regra de amostragem"
    if regra["tipo"] == "todas":
        return total_guias, "Crítica integral — auditar todas"
    if regra["tipo"] == "percentual":
        if total_procs < regra["minimo_procs"]:
            return total_guias, f"Menos de {regra['minimo_procs']} procs — auditar todas"
        n = max(1, round(total_guias * regra["pct"]))
        minimo_amostra = regra.get("minimo_amostra")
        if minimo_amostra:
            n = min(max(n, minimo_amostra), total_guias)
        pct = int(regra["pct"] * 100)
        sufixo = f" (mín. {minimo_amostra})" if minimo_amostra else ""
        return n, f"{pct}% das guias{sufixo} — auditar {n} de {total_guias}"
    return total_guias, ""


def sortear_amostra(df_guias: pd.DataFrame, n: int, seed: int) -> pd.DataFrame:
    if n >= len(df_guias):
        return df_guias.copy()
    return df_guias.sample(n=n, random_state=seed).sort_index()


def guias_com_proc_critico(df_esp_guias: pd.DataFrame, procedimentos_criticos: set) -> pd.DataFrame:
    """Filtra as guias (de uma especialidade fora de REGRAS_AMOSTRAGEM) que
    contêm pelo menos um procedimento cadastrado como crítico. Usado pra
    decidir se a especialidade sobe na lista e pra montar a "Sugestão de
    amostra" só com essas guias."""
    if df_esp_guias.empty or not procedimentos_criticos:
        return df_esp_guias.iloc[0:0]
    tem_critico = df_esp_guias["Procedimentos"].apply(
        lambda p: bool({c.strip() for c in p.split(",") if c.strip()} & procedimentos_criticos)
    )
    return df_esp_guias[tem_critico]


@st.cache_data(ttl=300)
def carregar_processos_turso() -> list:
    """Lista agregada de processos do mês (Turso) -- cacheada por 5min,
    mesmo padrão de carregar_dados_atuais() (REL5201), pra não reconsultar
    a cada rerun da tela."""
    from shared.database import DatabaseManager
    db = DatabaseManager()
    return db.listar_processos_agregado()


def montar_lista_processos_mes(
    processos_turso: list, df_rel5201: pd.DataFrame, procedimentos_criticos: set,
) -> pd.DataFrame:
    """Uma linha por processo do mês, cruzando a base IA (Turso, agregada
    por DatabaseManager.listar_processos_agregado) com o REL5201 já
    carregado -- pronta pra exibir/filtrar na lista de processos da
    Amostragem.

    % Liberação IA / Total de Guias / Procedimentos vêm do REL5201 -- é a
    contagem oficial do sistema, mais confiável que recalcular a partir da
    planilha da base IA (só um snapshot mensal, pode ficar defasado).
    Especialidades e crítica vêm do Turso -- o REL5201 não detalha por
    guia, só o total do processo.

    % Liberação IA = LIBERADOS_IA / (LIBERADOS_IA + NAO_LIBERADOS_IA) -- é
    por PROCEDIMENTO avaliado pela IA, não por guia (confirmado nos dados:
    a soma dos dois bate com QT_PROCEDIMENTO na maioria dos processos, quase
    nunca com QT_GUIAS).

    Status/Execução também vêm do REL5201 (STATUS_LABELS/valores normalizados
    já usados em Produtividade -- APP/MISTO/N_APP)."""
    from core.relatorio_5201 import STATUS_LABELS

    colunas_finais = [
        "Processo", "Status", "Execução", "% Liberação IA", "% Biometria", "Total de Guias",
        "Procedimentos", "Especialidades", "Crítica",
    ]
    if not processos_turso:
        return pd.DataFrame(columns=colunas_finais)

    _, ordem_criticas = carregar_regras_amostragem_cache()
    df_turso = pd.DataFrame(processos_turso)
    especialidades_lista = df_turso["especialidades"].fillna("").apply(
        lambda s: sorted({e.strip() for e in s.split(",") if e.strip()})
    )
    procedimentos_lista = df_turso["procedimentos"].fillna("").apply(
        lambda s: {p.strip() for p in s.split(",") if p.strip()}
    )
    tem_especialidade_critica = especialidades_lista.apply(
        lambda esps: any(_norm(e) in ordem_criticas for e in esps)
    )
    tem_procedimento_critico = procedimentos_lista.apply(
        lambda procs: bool(procs & procedimentos_criticos)
    )
    df_turso["Crítica"] = tem_especialidade_critica | tem_procedimento_critico
    df_turso["Especialidades"] = especialidades_lista.apply(lambda l: ", ".join(l))

    # % de procedimentos com biometria (mesma conta do card por-guia:
    # com-biometria / total, arredondado) -- em branco quando nenhum item do
    # processo tem operador gravado ainda (import antigo, sem essa coluna).
    itens_biometria = pd.to_numeric(df_turso.get("itens_biometria"), errors="coerce")
    itens_com_operador = pd.to_numeric(df_turso.get("itens_com_operador"), errors="coerce")
    total_itens = pd.to_numeric(df_turso.get("total_itens"), errors="coerce")
    pct_biometria = (itens_biometria / total_itens * 100).where(itens_com_operador > 0)
    df_turso["% Biometria"] = pct_biometria.round(1)

    campos_rel5201 = [
        "ORDEM", "STATUS", "EXECUCAO", "QT_GUIAS", "QT_PROCEDIMENTO",
        "QUANTIDADE_LIBERADOS_IA", "QUANTIDADE_NAO_LIBERADOS_IA",
    ]
    campos_disponiveis = [c for c in campos_rel5201 if c in df_rel5201.columns]
    if "ORDEM" in campos_disponiveis:
        df_rel = df_rel5201[campos_disponiveis].drop_duplicates(subset="ORDEM")
    else:
        df_rel = pd.DataFrame(columns=campos_rel5201)

    df_final = df_turso.merge(df_rel, left_on="nu_ordem", right_on="ORDEM", how="left")
    df_final["Processo"] = df_final["nu_ordem"]

    def _coluna_num(nome: str) -> pd.Series:
        # df_final.get(nome) devolve None (não uma Series) quando a coluna
        # não existe -- o REL5201 já importado antes dessas colunas serem
        # capturadas não as tem. Sem isso, contas tipo liberados+nao_liberados
        # quebram com TypeError em vez de só ficar em branco na tela.
        if nome in df_final.columns:
            return df_final[nome]
        return pd.Series(pd.NA, index=df_final.index, dtype="Int64")

    df_final["Total de Guias"] = _coluna_num("QT_GUIAS")
    df_final["Procedimentos"] = _coluna_num("QT_PROCEDIMENTO")

    liberados = _coluna_num("QUANTIDADE_LIBERADOS_IA")
    nao_liberados = _coluna_num("QUANTIDADE_NAO_LIBERADOS_IA")
    total_avaliado = liberados + nao_liberados
    pct_liberacao = (liberados / total_avaliado * 100).where(total_avaliado > 0)
    df_final["% Liberação IA"] = pct_liberacao.round(1)

    if "STATUS" in df_final.columns:
        df_final["Status"] = df_final["STATUS"].apply(lambda s: STATUS_LABELS.get(s, s) if s else None)
    else:
        df_final["Status"] = None

    if "EXECUCAO" in df_final.columns:
        df_final["Execução"] = df_final["EXECUCAO"].replace("", None)
    else:
        df_final["Execução"] = None

    # Processos sem nenhum match no REL5201 (Status vazio) vão pro fim da
    # lista -- não têm status/auditor/% pra mostrar, então atrapalham menos
    # lá embaixo do que misturados com o resto.
    sem_match = df_final["Status"].isna()
    return (
        df_final[colunas_finais]
        .assign(_sem_match=sem_match)
        .sort_values(["_sem_match", "Processo"])
        .drop(columns="_sem_match")
        .reset_index(drop=True)
    )


def _guia_tem_proc_prioritario(procs_str: str, procs_prioridade_normal: set) -> bool:
    """True se a guia contém pelo menos um procedimento fora da lista de
    prioridade normal (ou seja, um procedimento de alta prioridade que força
    inclusão garantida na amostra, sem depender do sorteio)."""
    procs = {p.strip() for p in procs_str.split(",") if p.strip()}
    return bool(procs - procs_prioridade_normal)


def marcar_amostra(df_esp_guias: pd.DataFrame, especialidade: str,
                   df_esp_procs_brutos: pd.DataFrame, seed: int) -> pd.DataFrame:
    """Retorna DataFrame de guias com a amostra sugerida.

    Fluxo:
      1. Aplica a regra da especialidade sobre o TOTAL de procs (não separa
         por prioridade de procedimento nesse passo). Se cai em 'auditar
         todas' (integral, ou < mínimo de procs), retorna tudo.
      2. Caso contrário, o tamanho da amostra = % das guias da especialidade
         (ex.: 30% pra CIRURGIA). Dentro desse tamanho, prioriza guias com
         procs de alta prioridade:
           - Se guias prioritárias ≤ 5: entram todas + completa com sorteio
             das normais até bater o tamanho da amostra.
           - Se guias prioritárias > 5: 50% da amostra vira prioritária
             sorteada + 50% vira normal sorteada.

    Coluna 'Motivo' é mantida por compatibilidade (dropada antes de renderizar).
    """
    # Guias sem nenhum procedimento remanescente (todos os procedimentos
    # foram marcados como ignorados) não têm o que auditar — ficam de fora
    # do sorteio/regra de amostra, mas continuam aparecendo na tabela
    # completa (fora desta função, que só recebe as elegíveis pra amostra).
    if "Qtde_procs" in df_esp_guias.columns:
        df_esp_guias = df_esp_guias[df_esp_guias["Qtde_procs"] > 0]

    if df_esp_guias.empty:
        return df_esp_guias.assign(Motivo=[])

    n_total_guias = len(df_esp_guias)
    total_procs_esp = int(df_esp_procs_brutos["Qtde"].sum())
    procs_prioridade_normal = PROCS_PRIORIDADE_NORMAL.get(_norm(especialidade))
    regras, _ = carregar_regras_amostragem_cache()
    regra = regras.get(_norm(especialidade), {})

    # Caminho 1: sem lista de prioridade normal OU regra que manda auditar tudo.
    def _todas(motivo=""):
        amostra = df_esp_guias.copy()
        amostra["Motivo"] = motivo
        return amostra

    if procs_prioridade_normal is None or not regra:
        n, _ = calcular_amostra(especialidade, total_procs_esp, n_total_guias, regras=regras)
        amostra = sortear_amostra(df_esp_guias, n, seed=seed).copy()
        amostra["Motivo"] = ""
        return amostra

    if regra.get("tipo") == "todas":
        return _todas("")

    if regra.get("tipo") == "percentual":
        # Gatilho <N procs = auditar todas — avaliado sobre o TOTAL da
        # especialidade, sem separar prioritárias de normais.
        if total_procs_esp < regra["minimo_procs"]:
            return _todas("")

        pct = regra["pct"]
        tamanho_amostra = max(1, round(n_total_guias * pct))
        minimo_amostra = regra.get("minimo_amostra")
        if minimo_amostra:
            # Nunca exige mais amostra do que existe de guias na especialidade.
            tamanho_amostra = min(max(tamanho_amostra, minimo_amostra), n_total_guias)

        df = df_esp_guias.copy()
        df["_prioritaria"] = df["Procedimentos"].apply(
            lambda p: _guia_tem_proc_prioritario(p, procs_prioridade_normal)
        )
        df_prioritarias = df[df["_prioritaria"]].drop(columns=["_prioritaria"]).copy()
        df_normais = df[~df["_prioritaria"]].drop(columns=["_prioritaria"]).copy()
        n_prior = len(df_prioritarias)
        n_norm = len(df_normais)

        if n_prior <= 5:
            # Todas as prioritárias entram (limitadas ao tamanho da amostra) e
            # o restante da amostra é preenchido com sorteio das normais.
            n_prior_amostra = min(n_prior, tamanho_amostra)
            n_norm_amostra = min(tamanho_amostra - n_prior_amostra, n_norm)
            # Se n_prior > tamanho_amostra, nem todas as prioritárias cabem —
            # sorteia quais entram em vez de estourar o tamanho declarado.
            df_prioritarias_final = (
                df_prioritarias if n_prior_amostra == n_prior
                else sortear_amostra(df_prioritarias, n_prior_amostra, seed=seed)
            )
        else:
            # Composição 50/50 dentro da amostra. Se um dos lados não tem
            # guias suficientes pra preencher a cota dele, o restante volta
            # pro outro lado -- senão a amostra final fica menor que
            # tamanho_amostra só porque um dos grupos estava vazio/pequeno
            # (ex.: todas as guias prioritárias, nenhuma normal).
            n_prior_amostra = min(tamanho_amostra // 2, n_prior)
            n_norm_amostra = min(tamanho_amostra - n_prior_amostra, n_norm)
            n_prior_amostra = min(n_prior_amostra + (tamanho_amostra - n_prior_amostra - n_norm_amostra), n_prior)
            df_prioritarias_final = sortear_amostra(df_prioritarias, n_prior_amostra, seed=seed)

        df_normais_final = sortear_amostra(df_normais, n_norm_amostra, seed=seed)

        df_prioritarias_final = df_prioritarias_final.copy()
        df_prioritarias_final["Motivo"] = ""
        df_normais_final = df_normais_final.copy()
        df_normais_final["Motivo"] = ""
        return pd.concat([df_prioritarias_final, df_normais_final], ignore_index=True)

    return _todas("")


def renderizar_tabela_guias(df_guias: pd.DataFrame, titulo_descritivo: str, objetivo: int,
                             guias_vistas: set = frozenset(), biometria_por_guia: dict = None,
                             imagem_por_guia: dict = None, guias_contagem: list = None):
    """Renderiza tabela HTML com NU_GUIA como botão clicável (copia ao clicar).

    `objetivo` é o tamanho de amostra requerido pela regra da especialidade
    (mostrado como denominador do contador).

    `guias_vistas`: NU_GUIA já marcados como auditados (vindos do Supabase —
    compartilhado entre qualquer um que abrir a mesma guia depois). Ao clicar,
    a marcação é gravada direto do navegador na tabela
    amostragem_guias_vistas (chave publicável do Supabase, sem vínculo a
    usuário por enquanto).

    `guias_contagem`: lista de NU_GUIA usada só pro NUMERADOR do contador,
    quando é diferente das linhas renderizadas nesta tabela -- caso da aba
    "Sugestão de amostra", que mostra só as N guias sorteadas mas cujo
    contador deve refletir TODAS as guias já vistas da especialidade (o
    auditor pode ter revisado guias fora da sugestão manualmente, e isso
    conta pra bater a meta). Pode passar de `objetivo` nesse caso -- é
    esperado, significa que a meta já foi superada. Se None, conta só as
    linhas desta tabela (comportamento padrão, usado em "Tabela completa" e
    "Sem Imagem").

    `biometria_por_guia`: {NU_GUIA: (qtd_com_biometria, qtd_total_itens,
    qtd_com_operador_gravado)}. Mostra "X/Y" na coluna BIOMETRIA; o ✓ só
    aparece quando X == Y (100% dos itens da guia atendidos via
    CONN_APPOD_NEW). Se nenhum item tiver operador gravado ainda (guia
    importada antes desta coluna existir), a célula fica em branco.

    `imagem_por_guia`: {NU_GUIA: (qtd_com_imagem, qtd_total)}, vindo da base
    de imagem (planilha própria, cruzada por NU_GUIA). Mesmo formato "X/Y"
    com ✓ só em 100%; guia ausente do dict (sem nenhum registro de imagem
    ainda) fica em branco.
    """
    biometria_por_guia = biometria_por_guia or {}
    imagem_por_guia = imagem_por_guia or {}
    supabase_url = st.secrets["supabase"]["url"].rstrip("/")
    supabase_key = st.secrets["supabase"]["key"]

    def _decodificar_sigla_requisito(req: str) -> str:
        """'RXIF' -> 'Radiografia Inicial e Final', preservando conectores
        '+' (precisa dos dois) e '/' (precisa de um dos dois) -- usado só
        pro tooltip (hover), a célula em si mostra a sigla curta."""
        partes = re.split(r"([+/])", req)
        saida = []
        for p in partes:
            if p == "+":
                saida.append(" + ")
            elif p == "/":
                saida.append(" ou ")
            else:
                saida.append(_DECODIFICA_SIGLA_REQUISITO.get(p, p))
        return "".join(saida)

    def _requisitos_guia(procedimentos_str: str) -> tuple[str, str]:
        """Requisito de documentação por guia, a partir do(s) código(s) de
        procedimento -- ver core/requisitos_procedimentos.py (297 códigos
        mapeados, fonte ainda não validada oficialmente). Se todos os
        procedimentos da guia pedem o MESMO requisito, mostra uma vez só;
        se pedem requisitos diferentes, mostra cada um com o código entre
        parênteses pra desambiguar (ex: "RXIF(5010), L(438)"). Procedimento
        sem requisito mapeado simplesmente não entra na lista -- não inventa
        nada pro que não temos dado. Retorna (texto_curto, tooltip_decodificado)."""
        codigos = [c.strip() for c in str(procedimentos_str).split(",") if c.strip()]
        pares = [(cod, REQUISITOS_POR_PROCEDIMENTO[cod]) for cod in codigos if cod in REQUISITOS_POR_PROCEDIMENTO]
        if not pares:
            return "", ""
        requisitos_unicos = {req for _, req in pares}
        if len(requisitos_unicos) == 1:
            req = next(iter(requisitos_unicos))
            return req, _decodificar_sigla_requisito(req)
        texto = ", ".join(f"{req}({cod})" for cod, req in pares)
        tooltip = "; ".join(f"{cod}: {_decodificar_sigla_requisito(req)}" for cod, req in pares)
        return texto, tooltip

    def _fracao_html(info):
        if info and info[1] > 0 and (len(info) < 3 or info[2] > 0):
            n_ok, n_total = info[0], info[1]
            if n_ok == n_total:
                return f"<td style='text-align:center'><span class='badge badge-ok'>{n_ok}/{n_total} ✓</span></td>"
            return f"<td style='text-align:center'><span class='badge badge-parcial'>{n_ok}/{n_total}</span></td>"
        return "<td style='text-align:center'><span class='badge badge-vazio'>—</span></td>"

    mostrar_motivo = "Motivo" in df_guias.columns
    linhas_html = []
    for _, row in df_guias.iterrows():
        guia = html.escape(str(row["NU_GUIA"]))
        procs_qtd = row.get("Procedimentos_qtd")
        if isinstance(procs_qtd, list) and procs_qtd:
            procs = ", ".join(
                f"{html.escape(str(cod))} <span class='qtd-badge'>×{qtd}</span>" if qtd > 1
                else html.escape(str(cod))
                for cod, qtd in procs_qtd
            )
        else:
            procs = html.escape(str(row["Procedimentos"]))
        requisitos_txt, requisitos_tooltip = _requisitos_guia(row["Procedimentos"])
        requisitos = html.escape(requisitos_txt)
        requisitos_html = (
            f"<td style='text-align:center'><span title='{html.escape(requisitos_tooltip)}'>{requisitos}</span></td>"
            if requisitos else "<td style='text-align:center'><span class='badge badge-vazio'>—</span></td>"
        )
        classe_vista = " vista" if str(row["NU_GUIA"]) in guias_vistas else ""
        biometria_html = _fracao_html(biometria_por_guia.get(str(row["NU_GUIA"])))
        imagem_html = _fracao_html(imagem_por_guia.get(str(row["NU_GUIA"])))
        motivo_html = ""
        if mostrar_motivo:
            motivo = html.escape(str(row.get("Motivo", "")))
            classe = "motivo-critica" if "crítica" in motivo.lower() or "critica" in motivo.lower() else "motivo-sorteio"
            motivo_html = f"<td class='{classe}'>{motivo}</td>"
        linhas_html.append(
            f"<tr>"
            f"<td><button class='copy-btn{classe_vista}' data-val='{guia}' title='Clique para copiar'>{guia}</button></td>"
            f"<td>{procs}</td>"
            f"{requisitos_html}"
            f"{biometria_html}"
            f"{imagem_html}"
            f"{motivo_html}"
            f"</tr>"
        )
    rows = "\n".join(linhas_html)
    th_motivo = "<th style='width: 18%'>Motivo</th>" if mostrar_motivo else ""

    # Contagem por especialidade inteira (quando guias_contagem é passado):
    # o numerador do contador soma qualquer guia da lista já vista, mesmo
    # que não esteja renderizada nesta tabela -- ver docstring acima.
    guias_contagem_json = json.dumps([str(g) for g in guias_contagem]) if guias_contagem is not None else "null"
    vistas_servidor_json = (
        json.dumps([str(g) for g in guias_contagem if str(g) in guias_vistas])
        if guias_contagem is not None else "[]"
    )

    html_tabela = f"""
    <style>
        body {{ color: #1f2937; background: transparent; margin: 0; }}
        .pbi-wrap {{ font-family: 'Source Sans Pro', sans-serif; color: inherit; }}
        .pbi-table {{ width: 100%; border-collapse: collapse; font-size: 14px; color: inherit; }}
        .pbi-table th, .pbi-table td {{
            padding: 7px 10px; text-align: left;
            border-bottom: 1px solid rgba(125,125,125,0.25);
            color: inherit;
        }}
        .pbi-table th {{
            background: #f0f2f6; font-weight: 600;
            position: sticky; top: 0; z-index: 1;
            box-shadow: 0 1px 0 rgba(125,125,125,0.35);
        }}
        .copy-btn {{
            background: transparent;
            border: 1px solid rgba(125,125,125,0.5);
            border-radius: 4px;
            padding: 3px 10px;
            cursor: pointer;
            font-family: ui-monospace, 'Cascadia Mono', Menlo, monospace;
            font-size: 13px;
            color: inherit;
        }}
        .copy-btn:hover {{ background: rgba(125,125,125,0.15); border-color: rgba(125,125,125,0.8); }}
        .copy-btn.vista {{
            background: rgba(46, 125, 50, 0.18);
            border-color: rgba(76, 175, 80, 0.65);
        }}
        .copy-btn.vista:hover {{ background: rgba(46, 125, 50, 0.32); }}
        .copy-btn.copied {{ background: #2e7d32; color: #fff; border-color: #43a047; }}
        .pbi-counter {{
            font-size: 12.5px;
            color: rgba(120,120,120,0.95);
            margin: 0 0 8px 2px;
            font-family: 'Source Sans Pro', sans-serif;
        }}
        .pbi-counter strong {{ color: rgba(76, 175, 80, 1); font-weight: 600; }}
        .pbi-counter.atingido strong {{ color: rgba(46, 125, 50, 1); }}
        .pbi-counter.atingido::after {{ content: ' ✓'; color: rgba(46, 125, 50, 1); font-weight: 600; }}
        .motivo-critica {{ color: #b45309; font-weight: 600; font-size: 12.5px; }}
        .motivo-sorteio {{ color: rgba(120,120,120,0.85); font-size: 12.5px; }}

        .badge {{
            display: inline-block;
            font-size: 12px;
            font-variant-numeric: tabular-nums;
            padding: 2px 9px;
            border-radius: 99px;
            border: 1px solid transparent;
            white-space: nowrap;
        }}
        .badge-ok {{ background: rgba(46, 125, 50, 0.18); color: #2e7d32; border-color: rgba(76, 175, 80, 0.45); }}
        .badge-parcial {{ background: rgba(180, 83, 9, 0.14); color: #b45309; border-color: rgba(180, 83, 9, 0.4); }}
        .badge-vazio {{ color: rgba(120,120,120,0.55); }}
        .qtd-badge {{
            display: inline-block;
            font-size: 11px;
            font-variant-numeric: tabular-nums;
            padding: 1px 6px;
            border-radius: 99px;
            background: rgba(56, 132, 224, 0.14);
            color: #2563a8;
            border: 1px solid rgba(56, 132, 224, 0.35);
            margin-left: 2px;
        }}

        @media (prefers-color-scheme: dark) {{
            body {{ color: #e6ecf5; }}
            .pbi-table th {{ background: #1c2230; box-shadow: 0 1px 0 rgba(255,255,255,0.15); }}
            .copy-btn {{ border-color: rgba(255,255,255,0.25); }}
            .copy-btn:hover {{ background: rgba(255,255,255,0.08); border-color: rgba(255,255,255,0.5); }}
            .copy-btn.vista {{
                background: rgba(76, 175, 80, 0.22);
                border-color: rgba(102, 187, 106, 0.75);
            }}
            .copy-btn.vista:hover {{ background: rgba(76, 175, 80, 0.4); }}
            .motivo-critica {{ color: #fbbf24; }}
            .badge-ok {{ background: rgba(76, 175, 80, 0.2); color: #7fd88a; border-color: rgba(102, 187, 106, 0.55); }}
            .badge-parcial {{ background: rgba(242, 169, 60, 0.16); color: #f2a93c; border-color: rgba(242, 169, 60, 0.45); }}
            .badge-vazio {{ color: rgba(255,255,255,0.35); }}
            .qtd-badge {{ background: rgba(96, 165, 250, 0.18); color: #93c5fd; border-color: rgba(96, 165, 250, 0.45); }}
            .motivo-sorteio {{ color: rgba(255,255,255,0.55); }}
        }}
    </style>
    <div class='pbi-wrap'>
        <div class='pbi-counter'><strong>0</strong> de {objetivo} analisado(s)</div>
        <table class='pbi-table'>
            <thead>
                <tr><th style='width: 25%'>NU_GUIA</th><th>Procedimentos</th><th style='width: 16%; text-align:center'>REQUISITOS</th><th style='width: 10%; text-align:center'>BIOMETRIA</th><th style='width: 10%; text-align:center'>IMAGEM</th>{th_motivo}</tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
    </div>
    <script>
        const PREFIX = 'amostragem_guia_vista_';
        const SUPABASE_URL = '{supabase_url}';
        const SUPABASE_KEY = '{supabase_key}';

        function marcarVistaNoServidor(nu_guia) {{
            fetch(`${{SUPABASE_URL}}/rest/v1/amostragem_guias_vistas`, {{
                method: 'POST',
                headers: {{
                    'apikey': SUPABASE_KEY,
                    'Authorization': `Bearer ${{SUPABASE_KEY}}`,
                    'Content-Type': 'application/json',
                    'Prefer': 'resolution=merge-duplicates,return=minimal',
                }},
                body: JSON.stringify({{ nu_guia: nu_guia }}),
            }}).catch(() => {{}});
        }}

        const OBJETIVO = {objetivo};
        const GUIAS_CONTAGEM = {guias_contagem_json};
        const VISTAS_SERVIDOR = {vistas_servidor_json};
        function atualizarContador() {{
            let vistos;
            if (GUIAS_CONTAGEM) {{
                const vistas = new Set(VISTAS_SERVIDOR);
                GUIAS_CONTAGEM.forEach(g => {{
                    if (localStorage.getItem(PREFIX + g) === '1') vistas.add(g);
                }});
                vistos = vistas.size;
            }} else {{
                vistos = document.querySelectorAll('.copy-btn.vista').length;
            }}
            const c = document.querySelector('.pbi-counter');
            if (!c) return;
            c.innerHTML = '<strong>' + vistos + '</strong> de ' + OBJETIVO + ' analisado(s)';
            if (vistos >= OBJETIVO) c.classList.add('atingido');
            else c.classList.remove('atingido');
        }}

        function aplicarEstadoVistas() {{
            document.querySelectorAll('.copy-btn').forEach(btn => {{
                const val = btn.getAttribute('data-val');
                if (localStorage.getItem(PREFIX + val) === '1') {{
                    btn.classList.add('vista');
                }}
            }});
            atualizarContador();
        }}

        aplicarEstadoVistas();

        // Sincroniza entre os dois iframes (tabela completa e amostra) na mesma janela
        window.addEventListener('storage', (e) => {{
            if (e.key && e.key.startsWith(PREFIX)) aplicarEstadoVistas();
        }});
        setInterval(aplicarEstadoVistas, 1500);

        document.querySelectorAll('.copy-btn').forEach(btn => {{
            btn.addEventListener('click', () => {{
                const val = btn.getAttribute('data-val');
                navigator.clipboard.writeText(val).then(() => {{
                    localStorage.setItem(PREFIX + val, '1');
                    marcarVistaNoServidor(val);
                    btn.classList.add('vista');
                    atualizarContador();
                    const orig = btn.innerText;
                    btn.innerText = '✓ ' + val;
                    btn.classList.add('copied');
                    setTimeout(() => {{
                        btn.innerText = orig;
                        btn.classList.remove('copied');
                    }}, 1100);
                }});
            }});
        }});
    </script>
    """
    altura = 82 + 36 * max(1, len(df_guias))
    components.html(html_tabela, height=min(altura, 540), scrolling=True)


def renderizar_botao_copiar_processo(processo) -> None:
    """Botão "Processo: X" que copia o número ao clicar -- mesmo padrão dos
    botões de NU_GUIA em renderizar_tabela_guias, só que sem o rastreio de
    "vista" (não se aplica ao processo em si). Existe porque é fácil perder
    o número do processo ao trocar de tela/copiar outra coisa por cima."""
    processo_esc = html.escape(str(processo))
    html_botao = f"""
    <style>
        body {{
            margin: 0; padding: 4px 0; background: transparent; color: #1f2937;
            font-family: 'Source Sans Pro', sans-serif;
            display: flex; align-items: center;
        }}
        .rotulo-processo {{ font-weight: 600; margin-right: 8px; }}
        .copy-btn-processo {{
            background: transparent;
            border: 1px solid rgba(125,125,125,0.5);
            border-radius: 4px;
            padding: 3px 10px;
            cursor: pointer;
            font-family: ui-monospace, 'Cascadia Mono', Menlo, monospace;
            font-size: 13px;
            color: inherit;
        }}
        .copy-btn-processo:hover {{ background: rgba(125,125,125,0.15); border-color: rgba(125,125,125,0.8); }}
        .copy-btn-processo.copied {{ background: #2e7d32; color: #fff; border-color: #43a047; }}
        @media (prefers-color-scheme: dark) {{
            body {{ color: #e6ecf5; }}
            .copy-btn-processo {{ border-color: rgba(255,255,255,0.25); }}
            .copy-btn-processo:hover {{ background: rgba(255,255,255,0.08); border-color: rgba(255,255,255,0.5); }}
        }}
    </style>
    <span class="rotulo-processo">Processo:</span>
    <button class="copy-btn-processo" id="btn-processo" title="Clique para copiar">{processo_esc}</button>
    <script>
        document.getElementById('btn-processo').addEventListener('click', () => {{
            navigator.clipboard.writeText('{processo_esc}').then(() => {{
                const btn = document.getElementById('btn-processo');
                const orig = btn.innerText;
                btn.innerText = '✓ copiado';
                btn.classList.add('copied');
                setTimeout(() => {{
                    btn.innerText = orig;
                    btn.classList.remove('copied');
                }}, 1100);
            }});
        }});
    </script>
    """
    components.html(html_botao, height=34)


def renderizar_resumo_especialidades(resumo: list, df: pd.DataFrame) -> None:
    """Tabela do Resumo em árvore (igual ao pivot do PowerBI que o time já usa:
    especialidade expansível mostrando quantidade por código de procedimento)
    -- ajuda a perceber se o prestador está concentrando pedidos num código
    específico. `resumo` é a lista de dicts (Especialidade, Guias únicas,
    Total de procs) já calculada por especialidade; `df` é
    o dataframe item-level (uma linha por procedimento, colunas Especialidade/
    CD_PROCEDIMENTO/Qtde) usado pra abrir a quebra por código."""
    linhas = []
    for i, item in enumerate(resumo):
        esp = item["Especialidade"]
        grupo_id = f"esp{i}"
        linhas.append(
            f"<tr class='linha-esp' onclick=\"toggleGrupo('{grupo_id}')\">"
            f"<td><span class='toggle-icon' id='icon-{grupo_id}'>+</span> {html.escape(str(esp))}</td>"
            f"<td style='text-align:center'>{item['Guias únicas']}</td>"
            f"<td style='text-align:center'>{item['Total de procs']}</td>"
            f"</tr>"
        )
        por_codigo = (
            df[df["Especialidade"] == esp].groupby("CD_PROCEDIMENTO")["Qtde"].sum()
            .sort_values(ascending=False)
        )
        for cod, qtd in por_codigo.items():
            linhas.append(
                f"<tr class='linha-codigo {grupo_id}' style='display:none'>"
                f"<td class='codigo-cell'>{html.escape(str(cod))}</td>"
                f"<td></td>"
                f"<td style='text-align:center'>{int(qtd)}</td>"
                f"</tr>"
            )
    rows = "\n".join(linhas)

    html_arvore = f"""
    <style>
        body {{ color: #1f2937; background: transparent; margin: 0; }}
        .pbi-wrap {{ font-family: 'Source Sans Pro', sans-serif; color: inherit; }}
        .pbi-table {{ width: 100%; border-collapse: collapse; font-size: 14px; color: inherit; }}
        .pbi-table th, .pbi-table td {{
            padding: 7px 10px; text-align: left;
            border-bottom: 1px solid rgba(125,125,125,0.25);
            color: inherit;
        }}
        .pbi-table th {{
            background: #f0f2f6; font-weight: 600;
            position: sticky; top: 0; z-index: 1;
            box-shadow: 0 1px 0 rgba(125,125,125,0.35);
        }}
        .linha-esp {{ cursor: pointer; font-weight: 600; }}
        .linha-esp:hover {{ background: rgba(125,125,125,0.08); }}
        .toggle-icon {{
            display: inline-block; width: 14px; text-align: center;
            color: rgba(120,120,120,0.9); font-weight: 700;
        }}
        .codigo-cell {{
            padding-left: 32px !important;
            color: rgba(120,120,120,0.95);
            font-weight: 400;
            font-family: ui-monospace, 'Cascadia Mono', Menlo, monospace;
        }}
        @media (prefers-color-scheme: dark) {{
            body {{ color: #e6ecf5; }}
            .pbi-table th {{ background: #1c2230; box-shadow: 0 1px 0 rgba(255,255,255,0.15); }}
            .linha-esp:hover {{ background: rgba(255,255,255,0.06); }}
            .codigo-cell {{ color: rgba(255,255,255,0.6); }}
        }}
    </style>
    <div class='pbi-wrap'>
        <table class='pbi-table'>
            <thead>
                <tr>
                    <th>Especialidade</th>
                    <th style='text-align:center'>Guias únicas</th>
                    <th style='text-align:center'>Total de procs</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
    </div>
    <script>
        function toggleGrupo(id) {{
            const icon = document.getElementById('icon-' + id);
            const abrir = icon.textContent === '+';
            document.querySelectorAll('.' + id).forEach(tr => {{
                tr.style.display = abrir ? 'table-row' : 'none';
            }});
            icon.textContent = abrir ? '−' : '+';
        }}
    </script>
    """
    # Altura reserva só as linhas visíveis de cara (especialidades fechadas)
    # -- content extra ao expandir rola dentro do iframe (scrolling=True),
    # em vez de reservar altura pra tudo aberto e sobrar espaço em branco
    # quando nada tá expandido.
    altura = 46 + 36 * len(resumo)
    components.html(html_arvore, height=min(altura, 420), scrolling=True)
